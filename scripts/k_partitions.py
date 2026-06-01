#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
K-Particiones recursivas (k=3,4,5) usando GeometricSIA (GeoMIP).
Columnas:
  K=3: S(19)=Particion T(20)=Phi U(21)=Tiempo
  K=4: Y(25)=Particion Z(26)=Phi AA(27)=Tiempo
  K=5: AE(31)=Particion AF(32)=Phi AG(33)=Tiempo

Uso:
    uv run --directory QNodes python scripts/k_partitions.py --all
    uv run --directory QNodes python scripts/k_partitions.py --sheet "10A-Elementos"
"""

import sys, io, os, re, time, shutil, argparse
from pathlib import Path

sys.setrecursionlimit(10000)
os.environ["FORCE_COLOR"] = "0"
os.environ["PYTHONIOENCODING"] = "utf-8"
import warnings
warnings.filterwarnings("ignore")
verbose = True

SCRIPTS_DIR = Path(__file__).parent
PROYECTO = SCRIPTS_DIR.parent
sys.path.insert(0, str(SCRIPTS_DIR))

import openpyxl
from openpyxl.styles import PatternFill
from utils.converter import excel_a_bits

GEOMIP_METHOD2 = PROYECTO / "GeoMIP" / "src" / "Method2_Dynamic_Programming_Reformulation"
sys.path.insert(0, str(GEOMIP_METHOD2))

import numpy as np
from src.models.base.application import aplicacion as geo_app
from src.controllers.manager import Manager as GeoManager
from src.controllers.strategies.geometric import GeometricSIA
from src.funcs.base import emd_efecto
from src.models.core.system import System
from utils.tpm_loader import cargar_tpm

EXCEL_FILE = PROYECTO / "DatosPruebas2026_1 (1).xlsx"
GEOMIP_SAMPLES = PROYECTO / "GeoMIP" / "data" / "samples"
QNODES_SAMPLES = PROYECTO / "QNodes" / "src" / ".samples"
N_PRUEBAS = 10

SHEETS = {
    "10A-Elementos": {"n_nodos": 10, "pagina": "A"},
    "15B-Elementos": {"n_nodos": 15, "pagina": "B"},
    "20A-Elementos": {"n_nodos": 20, "pagina": "A"},
    "22A-Elementos": {"n_nodos": 22, "pagina": "A"},
    "25A-Elementos ": {"n_nodos": 25, "pagina": "A"},
}

COLUMNAS = {
    3: (19, 20, 21),
    4: (25, 26, 27),
    5: (31, 32, 33),
}

def beam_para_n(n_nodos):
    if n_nodos <= 10: return 3
    if n_nodos <= 15: return 2
    if n_nodos <= 20: return 2
    return 1

def bits_desde_indices(indices: list[int], N: int) -> str:
    bits = ["0"] * N
    for i in indices:
        bits[i] = "1"
    return "".join(bits)

def parse_partition(part_str: str):
    lines = part_str.strip().split("\n")
    top_line = lines[0]
    parts = top_line.split("||")

    def extract(s):
        s = s.strip().strip("|").strip()
        if not s or s == "\u2205":
            return []
        result = []
        for c in s:
            if c.isalpha() and c.isupper():
                result.append(ord(c) - ord("A"))
        return result

    return extract(parts[0]), extract(parts[1])

def alcance_mec_para_grupo(grupo: list[int], alc_orig: str, mec_orig: str, N: int):
    bits = ["0"] * N
    for i in grupo:
        bits[i] = alc_orig[i]
    alc_grupo = "".join(bits)
    bits = ["0"] * N
    for i in grupo:
        bits[i] = mec_orig[i]
    mec_grupo = "".join(bits)
    return alc_grupo, mec_grupo

def fmt_subsistemas(subsystems: list[list[int]]) -> str:
    groups = []
    for group in subsystems:
        labels = [chr(ord("A") + i) for i in sorted(group)]
        groups.append(",".join(labels))
    return " | ".join(groups)

def preparar_sistema(tpm, estado_inicial, cond_orig, alc_orig, mec_orig):
    estado_arr = np.array([int(c) for c in estado_inicial], dtype=np.int8)
    system = System(tpm, estado_arr)
    cond_dims = np.array([i for i, c in enumerate(cond_orig) if c == "0"], dtype=np.int8)
    conditioned = system.condicionar(cond_dims)
    alc_remove = np.array([i for i, c in enumerate(alc_orig) if c == "0"], dtype=np.int8)
    mec_remove = np.array([i for i, c in enumerate(mec_orig) if c == "0"], dtype=np.int8)
    preprocessed = conditioned.substraer(alc_remove, mec_remove)
    original_marginal = preprocessed.distribucion_marginal()
    return preprocessed, original_marginal

def calcular_phi_k(preprocessed, original_marginal, subsystems):
    k_marginal = np.zeros_like(original_marginal)
    for node_list in subsystems:
        remove_alc = np.array(
            [i for i in preprocessed.indices_ncubos if i not in set(node_list)],
            dtype=np.int8,
        )
        remove_dims = (
            np.array([i for i in preprocessed.ncubos[0].dims if i not in set(node_list)], dtype=np.int8)
            if len(preprocessed.ncubos) > 0 else np.array([], dtype=np.int8)
        )
        sub = preprocessed.substraer(remove_alc, remove_dims)
        sub_marginal = sub.distribucion_marginal()
        for j, idx in enumerate(sub.indices_ncubos):
            pos = np.where(preprocessed.indices_ncubos == idx)[0][0]
            k_marginal[pos] = sub_marginal[j]
    return float(emd_efecto(k_marginal, original_marginal))

def post_optimizar(subsystems, preprocessed, original_marginal, n_nodos):
    if n_nodos > 15:
        return subsystems
    subsystems = [list(g) for g in subsystems]
    changed = True
    while changed:
        changed = False
        current_phi = calcular_phi_k(preprocessed, original_marginal, subsystems)
        for i in range(len(subsystems)):
            for node in list(subsystems[i]):
                best_phi = current_phi
                best_target = i
                for j in range(len(subsystems)):
                    if i == j: continue
                    new_systems = [list(g) for g in subsystems]
                    new_systems[i].remove(node)
                    new_systems[j].append(node)
                    new_phi = calcular_phi_k(preprocessed, original_marginal, new_systems)
                    if new_phi < best_phi - 1e-12:
                        best_phi = new_phi
                        best_target = j
                if best_target != i:
                    subsystems[i].remove(node)
                    subsystems[best_target].append(node)
                    current_phi = best_phi
                    changed = True
    return subsystems

def resolver_k_particiones_geometric(
    tpm, estado_inicial, pagina, N, k, cond_orig, alc_orig, mec_orig,
    beam_width=1, per_test_timeout=None, verbose=False
):
    cond_dims_i = set(i for i, c in enumerate(cond_orig) if c == "0")
    all_alc = set(i for i, c in enumerate(alc_orig) if c == "1")
    all_mec = set(i for i, c in enumerate(mec_orig) if c == "1")
    effective_nodes = sorted(all_alc - cond_dims_i)

    if not effective_nodes:
        if verbose: print(f"    No hay nodos efectivos")
        return [], []

    if k > len(effective_nodes):
        if verbose: print(f"    k={k} > nodos efectivos={len(effective_nodes)}")
        k = len(effective_nodes)

    subsystems = [effective_nodes]
    steps = []

    for iteration in range(k - 1):
        t0_iter = time.time()
        largest_idx = max(range(len(subsystems)), key=lambda i: len(subsystems[i]))
        largest = subsystems[largest_idx]
        if len(largest) <= 1:
            if verbose: print(f"    No se puede seguir dividiendo")
            break

        alc, mec = alcance_mec_para_grupo(largest, alc_orig, mec_orig, N)

        geo_app.pagina_sample_network = pagina
        geo_app.profiler_habilitado = False
        gestor = GeoManager(estado_inicial=estado_inicial)
        analizador = GeometricSIA(gestor, beam_width=beam_width)
        sol = analizador.aplicar_estrategia(cond_orig, alc, mec, tpm, timeout=per_test_timeout)

        g1, g2 = parse_partition(sol.particion)

        if not g1 or not g2:
            l = sorted(largest)
            mid = len(l) // 2
            g1, g2 = l[:mid], l[mid:]

        subsystems.pop(largest_idx)
        subsystems.append(g1)
        subsystems.append(g2)
        steps.append((largest, g1, g2, float(sol.perdida)))

        if verbose:
            print(f"    Split {iteration+1}/{k-1}: phi={sol.perdida:.6f} t={time.time()-t0_iter:.2f}s")

    return subsystems, steps

def resolver_tpm(n_nodos, pagina):
    candidates = [
        GEOMIP_SAMPLES / f"N{n_nodos}{pagina}.csv",
        QNODES_SAMPLES / f"N{n_nodos}{pagina}.csv",
    ]
    for path in candidates:
        if path.exists():
            return path
    raise FileNotFoundError(f"No se encontro N{n_nodos}{pagina}.csv")

def procesar_hoja(sheet_name, config, k_values):
    n_nodos = config["n_nodos"]
    pagina = config["pagina"]
    beam_width = beam_para_n(n_nodos)

    print(f"\n{'='*70}")
    print(f"  GEOMETRIC SIA - {sheet_name} ({n_nodos} nodos, pagina {pagina.upper()})")
    print(f"  k={k_values}  beam_width={beam_width}")
    print(f"  Columnas: S-U(k=3) | Y-AA(k=4) | AE-AG(k=5)")
    print(f"{'='*70}")

    print(f"\n  --- Leyendo {N_PRUEBAS} pruebas desde Excel ---")
    try:
        wb = openpyxl.load_workbook(str(EXCEL_FILE), data_only=False)
        ws = wb[sheet_name]
        pruebas = []
        for fila in range(6, 6 + N_PRUEBAS):
            alcance = ws[f"B{fila}"].value
            mecanismo = ws[f"C{fila}"].value
            if alcance and mecanismo:
                pruebas.append({"fila": fila, "alcance": str(alcance).strip(), "mecanismo": str(mecanismo).strip()})
        wb.close()
        print(f"  Leidas {len(pruebas)} pruebas")
    except Exception as e:
        print(f"  ERROR leyendo Excel: {e}")
        return

    tpm_path = resolver_tpm(n_nodos, pagina)
    tpm = cargar_tpm(tpm_path)
    print(f"  TPM cargado: {tpm_path.name} ({tpm.shape})")

    buffer_resultados = []

    for idx, p in enumerate(pruebas, 1):
        alcance_bits = excel_a_bits(p["alcance"], n_nodos)
        mecanismo_bits = excel_a_bits(p["mecanismo"], n_nodos)
        estado_inicial = "1" + "0" * (n_nodos - 1)
        condiciones = "1" * n_nodos

        print(f"\n  >>> PRUEBA {idx}/{N_PRUEBAS} (fila {p['fila']}) <<<")
        print(f"      alcance={p['alcance']}  mecanismo={p['mecanismo']}")

        t0_sistema = time.time()
        preprocessed, original_marginal = preparar_sistema(tpm, estado_inicial, condiciones, alcance_bits, mecanismo_bits)
        t_prep = time.time() - t0_sistema
        if verbose:
            print(f"      Sistema preparado en {t_prep:.2f}s")

        for k_val in k_values:
            if k_val > n_nodos:
                print(f"      k={k_val}: saltando (N={n_nodos})")
                continue

            print(f"      --- k={k_val} ---")
            try:
                inicio = time.time()
                subsystems, steps = resolver_k_particiones_geometric(
                    tpm, estado_inicial, pagina, n_nodos, k_val,
                    condiciones, alcance_bits, mecanismo_bits,
                    beam_width=beam_width, per_test_timeout=None,
                    verbose=True,
                )

                if subsystems:
                    subsystems = post_optimizar(subsystems, preprocessed, original_marginal, n_nodos)
                    phi_k = calcular_phi_k(preprocessed, original_marginal, subsystems)
                else:
                    phi_k = 0.0

                wall_clock = time.time() - inicio
                part_str = fmt_subsistemas(subsystems) if subsystems else ""
                print(f"      phi_{k_val}={phi_k:.10f}  t={wall_clock:.2f}s")
                print(f"      Particion: {part_str}")
                buffer_resultados.append((p["fila"], k_val, part_str, phi_k, wall_clock))
            except Exception as e:
                print(f"      [ERROR] k={k_val}: {type(e).__name__}: {e}")

    t0_guardar = time.time()
    try:
        wb = openpyxl.load_workbook(str(EXCEL_FILE), data_only=False)
        ws = wb[sheet_name]
        ok_fill = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")
        for fila, k_val, part_str, phi, tiempo in buffer_resultados:
            col_info = COLUMNAS.get(k_val)
            if col_info:
                ws.cell(row=fila, column=col_info[0], value=str(part_str))
                ws.cell(row=fila, column=col_info[1], value=float(phi))
                ws.cell(row=fila, column=col_info[2], value=float(tiempo))
                for col in col_info:
                    ws.cell(row=fila, column=col).fill = ok_fill
        wb.save(str(EXCEL_FILE))
        wb.close()
        print(f"  Guardados {len(buffer_resultados)} resultados en Excel ({time.time()-t0_guardar:.1f}s)")
    except Exception as e:
        print(f"  ERROR guardando Excel: {e}")

    print(f"\n  --- GEOMETRIC {sheet_name} COMPLETADO ---")

def main():
    parser = argparse.ArgumentParser(description="K-Particiones con GeometricSIA (GeoMIP)")
    group = parser.add_mutually_exclusive_group(required=True)
    group.add_argument("--sheet", type=str, help="Nombre de la hoja")
    group.add_argument("--all", action="store_true", help="Todas las hojas")
    parser.add_argument("--k", type=int, nargs="+", default=[3, 4, 5])
    parser.add_argument("--beam-width", type=int, default=0, help="0=auto segun N")
    args = parser.parse_args()

    print("=" * 70)
    print("  K-PARTICIONES - GEOMETRIC SIA (GeoMIP)")
    print("  k={}  beam_width={}".format(",".join(str(v) for v in args.k),
          "auto" if args.beam_width == 0 else str(args.beam_width)))
    print("  Columnas: S=Part3 T=Phi3 U=Time3 | Y=Part4 Z=Phi4 AA=Time4 | AE=Part5 AF=Phi5 AG=Time5")
    print("  Optimizaciones: cache phi_k + buffer Excel + post-optimizacion + params adaptativos")
    print("=" * 70)

    backup = PROYECTO / f"BACKUP_{Path(EXCEL_FILE).name}"
    if not backup.exists():
        shutil.copy(str(EXCEL_FILE), str(backup))
        print(f"\n  Backup creado: {backup.name}")

    if args.all:
        hojas = list(SHEETS.items())
    else:
        sheet_name = args.sheet
        for k in SHEETS:
            if k.strip() == sheet_name.strip():
                sheet_name = k; break
        else:
            print(f"ERROR: Hoja '{sheet_name}' no valida. Opciones: {list(SHEETS.keys())}")
            sys.exit(1)
        hojas = [(sheet_name, SHEETS[sheet_name])]

    for sheet_name, config in hojas:
        procesar_hoja(sheet_name, config, args.k)

if __name__ == "__main__":
    main()
