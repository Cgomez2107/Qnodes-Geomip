#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
PRUEBAS QNODES - VERSION v5
Escribe a archivo temporal, verifica, y luego reemplaza.
"""

import sys
import os
import io
import time
import shutil
from pathlib import Path

sys.stdout = io.TextIOWrapper(io.BufferedWriter(sys.stdout.buffer), encoding='utf-8', errors='replace')

PROYECTO = Path(r"c:\Users\carli\Documents\U. Caldas\Algoritmos\Proyecto\Proyecto_Analisis_2026_1")
EXCEL_FILE = PROYECTO / "DatosPruebas2026_1 (1).xlsx"
EXCEL_TEMP = PROYECTO / "DATOS_TEMP_NUEVO.xlsx"
EXCEL_BACKUP = PROYECTO / "DATOS_BACKUP.xlsx"

sys.path.insert(0, str(PROYECTO / "scripts"))

from utils.converter import excel_a_bits
from utils.parser import parsear_salida_qnodes
from utils.excel_handler import leer_pruebas

import subprocess
import re


def main():
    print("\n" + "="*70)
    print(" PRUEBAS QNODES - 10A-Elementos (v5)")
    print("="*70 + "\n")
    
    # 1. Leer pruebas
    print("[PASO 1] Leyendo pruebas...")
    try:
        pruebas = leer_pruebas(str(EXCEL_FILE), "10A-Elementos", 10)
        print(f"  ✓ {len(pruebas)} pruebas\n")
    except Exception as e:
        print(f"  ✗ {e}\n")
        return False
    
    # 2. Ejecutar pruebas
    print("[PASO 2] Ejecutando QNodes...\n")
    resultados = {}
    
    for idx, prueba in enumerate(pruebas, 1):
        num = prueba["numero"]
        fila = prueba["fila"]
        alcance_excel = prueba["alcance"]
        mecanismo_excel = prueba["mecanismo"]
        
        alcance_bits = excel_a_bits(alcance_excel, 10)
        mecanismo_bits = excel_a_bits(mecanismo_excel, 10)
        
        print(f"  Prueba {idx}/10: ", end="", flush=True)
        
        try:
            # Actualizar main.py
            main_py = PROYECTO / "QNodes" / "src" / "main.py"
            with open(main_py, 'r', encoding='utf-8') as f:
                content = f.read()
            
            content = re.sub(r'estado_inicial\s*=\s*["\'][\d]+["\']', 'estado_inicial = "1000000000"', content)
            content = re.sub(r'condiciones\s*=\s*["\'][\d]+["\']', 'condiciones = "1111111111"', content)
            content = re.sub(r'alcance\s*=\s*["\'][\d]+["\']', f'alcance = "{alcance_bits}"', content)
            content = re.sub(r'mecanismo\s*=\s*["\'][\d]+["\']', f'mecanismo = "{mecanismo_bits}"', content)
            
            with open(main_py, 'w', encoding='utf-8') as f:
                f.write(content)
            
            # Ejecutar QNodes
            result = subprocess.run(
                ["uv", "run", "exec.py"],
                cwd=str(PROYECTO / "QNodes"),
                capture_output=True,
                text=True,
                timeout=60,
                encoding='utf-8',
                errors='replace'
            )
            
            # Parsear
            output = result.stdout + result.stderr
            data = parsear_salida_qnodes(output)
            phi = data.get("phi")
            tiempo = data.get("tiempo")
            
            if phi is None or tiempo is None:
                print(f"✗")
                continue
            
            resultados[num] = {"fila": fila, "phi": phi, "tiempo": tiempo}
            print(f"✓")
            
        except Exception as e:
            print(f"✗")
    
    print()
    
    if not resultados:
        print("[ERROR] No hay resultados\n")
        return False
    
    # 3. Crear archivo temporal con datos
    print("[PASO 3] Creando archivo temporal con resultados...\n")
    try:
        import openpyxl
        
        # Copiar archivo original al temp
        print(f"  Copiando {EXCEL_FILE.name} -> {EXCEL_TEMP.name}...", end="", flush=True)
        shutil.copy(str(EXCEL_FILE), str(EXCEL_TEMP))
        print(" ✓")
        
        # Escribir en temp
        print(f"  Escribiendo datos en {EXCEL_TEMP.name}...", end="", flush=True)
        wb = openpyxl.load_workbook(str(EXCEL_TEMP))
        ws = wb["10A-Elementos"]
        
        for num, data in resultados.items():
            fila = data["fila"]
            phi = float(data["phi"])
            tiempo = float(data["tiempo"])
            ws[f"E{fila}"] = phi
            ws[f"F{fila}"] = tiempo
        
        wb.save(str(EXCEL_TEMP))
        wb.close()
        print(" ✓")
        
        # Verificar temp
        print(f"  Verificando {EXCEL_TEMP.name}...", end="", flush=True)
        wb = openpyxl.load_workbook(str(EXCEL_TEMP))
        ws = wb["10A-Elementos"]
        for fila in range(6, 16):
            phi = ws[f"E{fila}"].value
            tiempo = ws[f"F{fila}"].value
            if phi is None or tiempo is None:
                print(f" ✗ (fila {fila} vacía)")
                wb.close()
                return False
        wb.close()
        print(" ✓")
        
    except Exception as e:
        print(f" ✗ ({e})\n")
        return False
    
    # 4. Reemplazar archivo original
    print(f"\n[PASO 4] Reemplazando archivo original...\n")
    try:
        # Hacer backup
        print(f"  Backup: {EXCEL_FILE.name} -> {EXCEL_BACKUP.name}...", end="", flush=True)
        if EXCEL_BACKUP.exists():
            os.remove(str(EXCEL_BACKUP))
        shutil.copy(str(EXCEL_FILE), str(EXCEL_BACKUP))
        print(" ✓")
        
        # Esperar y reemplazar
        print(f"  Reemplazando archivo original...", end="", flush=True)
        for intento in range(5):
            try:
                os.remove(str(EXCEL_FILE))
                shutil.move(str(EXCEL_TEMP), str(EXCEL_FILE))
                print(" ✓")
                break
            except Exception as e:
                if intento < 4:
                    time.sleep(1)
                else:
                    raise
        
    except Exception as e:
        print(f" ✗ ({e})\n")
        print(f"  [INFO] Datos están en {EXCEL_TEMP}")
        return False
    
    # 5. Verificar archivo final
    print(f"\n[PASO 5] Verificando archivo final...\n")
    try:
        import openpyxl
        wb = openpyxl.load_workbook(str(EXCEL_FILE))
        ws = wb["10A-Elementos"]
        
        print(f"  Contenido de {EXCEL_FILE.name}:")
        for fila in range(6, 16):
            phi = ws[f"E{fila}"].value
            tiempo = ws[f"F{fila}"].value
            print(f"    Fila {fila}: φ={phi}, t={tiempo}")
        
        wb.close()
        
    except Exception as e:
        print(f"  Error verificando: {e}\n")
        return False
    
    print("\n" + "="*70)
    print(f" ✓ COMPLETADO: {len(resultados)} pruebas")
    print("="*70 + "\n")
    
    return True


if __name__ == "__main__":
    try:
        success = main()
        sys.exit(0 if success else 1)
    except Exception as e:
        print(f"\n[ERROR] {e}\n")
        sys.exit(1)
