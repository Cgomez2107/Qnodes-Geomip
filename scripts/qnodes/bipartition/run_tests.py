#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
PRUEBAS QNODES - VERSION FINAL v4
Completamente rediseñado con mejor manejo de errores.
"""

import sys
import os
import io
import time
from pathlib import Path

# UTF-8 setup
sys.stdout = io.TextIOWrapper(io.BufferedWriter(sys.stdout.buffer), encoding='utf-8', errors='replace')

PROYECTO = Path(r"c:\Users\carli\Documents\U. Caldas\Algoritmos\Proyecto\Proyecto_Analisis_2026_1")
EXCEL_FILE = PROYECTO / "DatosPruebas2026_1 (1).xlsx"

sys.path.insert(0, str(PROYECTO / "scripts"))

from utils.converter import excel_a_bits
from utils.parser import parsear_salida_qnodes
from utils.excel_handler import leer_pruebas

import subprocess
import re


def main():
    print("\n" + "="*70)
    print(" PRUEBAS QNODES - 10A-Elementos")
    print("="*70 + "\n")
    
    # 1. Leer pruebas
    print("[PASO 1] Leyendo pruebas del Excel...")
    try:
        pruebas = leer_pruebas(str(EXCEL_FILE), "10A-Elementos", 10)
        print(f"  ✓ {len(pruebas)} pruebas leídas\n")
    except Exception as e:
        print(f"  ✗ Error: {e}\n")
        return False
    
    # 2. Ejecutar pruebas
    print("[PASO 2] Ejecutando QNodes para cada prueba...\n")
    resultados = {}
    
    for idx, prueba in enumerate(pruebas, 1):
        num = prueba["numero"]
        fila = prueba["fila"]
        alcance_excel = prueba["alcance"]
        mecanismo_excel = prueba["mecanismo"]
        
        alcance_bits = excel_a_bits(alcance_excel, 10)
        mecanismo_bits = excel_a_bits(mecanismo_excel, 10)
        
        print(f"  Prueba {idx}/10: ", end="", flush=True)
        
        try:
            # Actualizar main.py
            main_py = PROYECTO / "QNodes" / "src" / "main.py"
            with open(main_py, 'r', encoding='utf-8') as f:
                content = f.read()
            
            content = re.sub(r'estado_inicial\s*=\s*["\'][\d]+["\']', 'estado_inicial = "1000000000"', content)
            content = re.sub(r'condiciones\s*=\s*["\'][\d]+["\']', 'condiciones = "1111111111"', content)
            content = re.sub(r'alcance\s*=\s*["\'][\d]+["\']', f'alcance = "{alcance_bits}"', content)
            content = re.sub(r'mecanismo\s*=\s*["\'][\d]+["\']', f'mecanismo = "{mecanismo_bits}"', content)
            
            with open(main_py, 'w', encoding='utf-8') as f:
                f.write(content)
            
            # Ejecutar QNodes
            result = subprocess.run(
                ["uv", "run", "exec.py"],
                cwd=str(PROYECTO / "QNodes"),
                capture_output=True,
                text=True,
                timeout=60,
                encoding='utf-8',
                errors='replace'
            )
            
            # Parsear
            output = result.stdout + result.stderr
            data = parsear_salida_qnodes(output)
            phi = data.get("phi")
            tiempo = data.get("tiempo")
            
            if phi is None or tiempo is None:
                print(f"✗ (parsing error)")
                continue
            
            resultados[num] = {"fila": fila, "phi": phi, "tiempo": tiempo}
            print(f"✓ (φ={phi:.4f}, t={tiempo:.4f}s)")
            
        except Exception as e:
            print(f"✗ ({type(e).__name__})")
    
    print()
    
    if not resultados:
        print("[ERROR] No hay resultados\n")
        return False
    
    print(f"[PASO 3] Escribiendo {len(resultados)} resultados en Excel...\n")
    
    # Escribir Excel
    try:
        import openpyxl
        
        wb = openpyxl.load_workbook(str(EXCEL_FILE))
        ws = wb["10A-Elementos"]
        
        for num, data in resultados.items():
            fila = data["fila"]
            phi = float(data["phi"])
            tiempo = float(data["tiempo"])
            
            ws[f"E{fila}"] = phi
            ws[f"F{fila}"] = tiempo
            print(f"  Fila {fila}: φ = {phi:.6f}, t = {tiempo:.4f}s")
        
        wb.save(str(EXCEL_FILE))
        wb.close()
        
        print(f"\n✓ Archivo guardado: {EXCEL_FILE}\n")
        
    except Exception as e:
        print(f"\n✗ Error escribiendo Excel: {e}\n")
        return False
    
    # Resumen
    print("="*70)
    print(f" COMPLETADO: {len(resultados)} pruebas exitosas")
    print("="*70 + "\n")
    
    return True


if __name__ == "__main__":
    try:
        success = main()
        sys.exit(0 if success else 1)
    except Exception as e:
        print(f"\n[ERROR FATAL] {e}\n")
        sys.exit(1)
