"""
Handler de Excel: Lee y escribe resultados en el archivo .xlsx
"""

import openpyxl
from openpyxl.styles import Font, PatternFill


def leer_pruebas(archivo_excel, sheet_name, n_pruebas=10):
    """
    Lee las pruebas de una hoja del excel.
    
    Retorna lista de dicts con:
    - numero: #Prueba (1-10)
    - alcance: Notación Excel
    - mecanismo: Notación Excel
    """
    
    try:
        wb = openpyxl.load_workbook(archivo_excel, data_only=False)
        ws = wb[sheet_name]
        
        pruebas = []
        
        # Las pruebas comienzan en fila 6 (fila 1-5 son headers)
        # Leer filas 6 a 6+n_pruebas
        numero_prueba = 1
        for fila in range(6, 6 + n_pruebas):
            alcance = ws[f"B{fila}"].value
            mecanismo = ws[f"C{fila}"].value
            
            if alcance and mecanismo:  # Solo si tiene valores
                pruebas.append({
                    "numero": numero_prueba,
                    "fila": fila,
                    "alcance": str(alcance).strip(),
                    "mecanismo": str(mecanismo).strip()
                })
                numero_prueba += 1
        
        wb.close()
        return pruebas
    
    except FileNotFoundError:
        print(f"ERROR: Archivo no encontrado: {archivo_excel}")
        return []
    except KeyError:
        print(f"ERROR: Hoja '{sheet_name}' no encontrada en el excel")
        return []


def escribir_resultado_qnodes(archivo_excel, sheet_name, fila, particion, phi, tiempo):
    """
    Escribe un resultado en el excel en la fila especificada.
    
    Columnas QNodes:
    - D: Partición
    - E: Pérdida (Φ)
    - F: Tiempo (s)
    """
    
    import time
    import gc
    import os
    import subprocess
    
    max_reintentos = 20
    
    for intento in range(max_reintentos):
        try:
            gc.collect()
            delay = min(0.1 * (2 ** (intento // 5)), 2.0)
            time.sleep(delay)
            
            wb = openpyxl.load_workbook(archivo_excel)
            ws = wb[sheet_name]
            
            # Escribir valores
            ws[f"D{fila}"] = str(particion) if particion else ""
            ws[f"E{fila}"] = float(phi) if phi is not None else ""
            ws[f"F{fila}"] = float(tiempo) if tiempo is not None else ""
            
            # Formatear con colores
            fill_color = PatternFill(start_color="FFB6D9EC", end_color="FFB6D9EC", fill_type="solid")
            for col in ["D", "E", "F"]:
                ws[f"{col}{fila}"].fill = fill_color
            
            wb.save(archivo_excel)
            wb.close()
            gc.collect()
            
            return True
        
        except PermissionError as e:
            if intento == max_reintentos - 1:
                try:
                    ps_cmd = f"""
                    $file = '{archivo_excel}'
                    [System.Runtime.Interop Services.Marshal]::ReleaseComObject([System.Runtime.InteropServices.Marshal]::GetActiveObject('Excel.Application')) | Out-Null
                    """
                    subprocess.run(["powershell", "-Command", ps_cmd], timeout=5, capture_output=True)
                except:
                    pass
                return False
        except Exception as e:
            if intento == max_reintentos - 1:
                print(f" [{type(e).__name__}]")
                return False


def escribir_resultado_geometric(archivo_excel, sheet_name, fila, particion, phi, tiempo):
    """
    Escribe un resultado en el excel para GeoMIP.
    
    Columnas Geometric: G=Partición, H=Pérdida, I=Tiempo
    """
    
    try:
        wb = openpyxl.load_workbook(archivo_excel)
        ws = wb[sheet_name]
        
        # Escribir valores
        ws[f"G{fila}"] = particion if particion else ""
        ws[f"H{fila}"] = float(phi) if phi else ""
        ws[f"I{fila}"] = float(tiempo) if tiempo else ""
        
        # Formatear con colores
        fill_color = PatternFill(start_color="FFC6E0B4", end_color="FFC6E0B4", fill_type="solid")
        for col in ["G", "H", "I"]:
            ws[f"{col}{fila}"].fill = fill_color
        
        wb.save(archivo_excel)
        wb.close()
        
        return True
    
    except Exception as e:
        return False


# Tests
if __name__ == "__main__":
    archivo = r"C:\Users\carli\Documents\U. Caldas\Algoritmos\Proyecto\Proyecto_Analisis_2026_1\DatosPruebas2026_1 (1).xlsx"
    
    # Leer pruebas
    pruebas = leer_pruebas(archivo, "10A-Elementos")
    for p in pruebas[:2]:
        print(f"Prueba {p['numero']}: alcance={p['alcance']}, mecanismo={p['mecanismo']}")
