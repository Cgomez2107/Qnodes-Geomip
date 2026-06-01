#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Extrae y analiza resultados de k-partitions: QNodes vs GeometricSIA
Lee del Excel y genera tablas + gráficas en reports/
"""

import sys, os, json
from pathlib import Path
sys.setrecursionlimit(10000)
os.environ["FORCE_COLOR"] = "0"

SCRIPTS_DIR = Path(__file__).parent
PROYECTO = SCRIPTS_DIR.parent
sys.path.insert(0, str(PROYECTO / "QNodes"))

import openpyxl
import numpy as np
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt

EXCEL_FILE = PROYECTO / "DatosPruebas2026_1 (1).xlsx"
REPORTS_DIR = PROYECTO / "reports"
REPORTS_DIR.mkdir(exist_ok=True)

SHEETS = ["10A-Elementos", "15B-Elementos", "20A-Elementos"]
COLUMNAS_Q = {3: (16,17,18), 4: (22,23,24), 5: (28,29,30)}
COLUMNAS_G = {3: (19,20,21), 4: (25,26,27), 5: (31,32,33)}

def extraer_datos():
    wb = openpyxl.load_workbook(str(EXCEL_FILE), data_only=True)
    datos = {}
    for sheet_name in SHEETS:
        ws = wb[sheet_name]
        tests = []
        for fila in range(6, 16):
            alcance = ws.cell(fila, 2).value or ""
            mecanismo = ws.cell(fila, 3).value or ""
            test = {"fila": fila, "alcance": str(alcance).strip(), "mecanismo": str(mecanismo).strip()}
            for k in [3,4,5]:
                for metodo, cols in [("Q", COLUMNAS_Q), ("G", COLUMNAS_G)]:
                    pv = ws.cell(fila, cols[k][0]).value
                    fv = ws.cell(fila, cols[k][1]).value
                    tv = ws.cell(fila, cols[k][2]).value
                    pref = f"{metodo}{k}"
                    test[f"{pref}_part"] = str(pv).strip() if pv is not None else ""
                    test[f"{pref}_phi"] = float(fv) if fv is not None else None
                    test[f"{pref}_time"] = float(tv) if tv is not None else None
            tests.append(test)
        datos[sheet_name] = tests
    wb.close()
    return datos

def resumen_por_hoja(datos):
    print("=" * 100)
    print(f"{'Hoja':<20} {'k':<5} {'Metodo':<12} {'phi_prom':<15} {'phi_best':<12} {'phi_worst':<12} {'Time_avg':<15} {'phi=0':<15}")
    print("=" * 100)
    for sheet in SHEETS:
        tests = datos[sheet]
        for k in [3,4,5]:
            for metodo, pref, nombre in [("Q", "Q", "QNodes"), ("G", "G", "GeometricSIA")]:
                phis = [t[f"{pref}{k}_phi"] for t in tests if t[f"{pref}{k}_phi"] is not None]
                times = [t[f"{pref}{k}_time"] for t in tests if t[f"{pref}{k}_time"] is not None]
                if not phis:
                    print(f"{sheet:<20} {k:<5} {nombre:<12} {'SIN DATOS':<54}")
                    continue
                phi_zeros = sum(1 for p in phis if abs(p) < 1e-9)
                print(f"{sheet:<20} {k:<5} {nombre:<12} {np.mean(phis):<15.10f} {min(phis):<12.10f} {max(phis):<12.10f} {np.mean(times):<15.2f}s {phi_zeros}/{len(phis)}")

def comparar_por_test(datos):
    print("\n" + "=" * 120)
    print(f"{'Hoja':<15} {'Test':<5} {'k':<5} {'phi_Q':<15} {'phi_G':<15} {'Dif':<10} {'Gana':<8} {'t_Q':<10} {'t_G':<10}")
    print("=" * 120)
    q_gana = 0; g_gana = 0; empates = 0
    for sheet in SHEETS:
        tests = datos[sheet]
        for k in [3,4,5]:
            for i, t in enumerate(tests, 1):
                pq = t[f"Q{k}_phi"]
                pg = t[f"G{k}_phi"]
                tq = t[f"Q{k}_time"]
                tg = t[f"G{k}_time"]
                if pq is None or pg is None:
                    print(f"{sheet:<15} {i:<5} {k:<5} {'N/A':<15} {'N/A':<15} {'N/A':<10} {'N/A':<8} {'N/A':<10} {'N/A':<10}")
                    continue
                diff = pq - pg
                if diff < -1e-9:
                    ganador = "QNodes"
                    q_gana += 1
                elif diff > 1e-9:
                    ganador = "GeoSIA"
                    g_gana += 1
                else:
                    ganador = "Empate"
                    empates += 1
                print(f"{sheet:<15} {i:<5} {k:<5} {pq:<15.10f} {pg:<15.10f} {diff:<+10.10f} {ganador:<8} {tq:<10.2f}s {tg:<10.2f}s")
    total = q_gana + g_gana + empates
    print(f"\n  QNodes gana: {q_gana}/{total} | GeometricSIA gana: {g_gana}/{total} | Empates: {empates}/{total}")

def graficar_comparacion(datos):
    for k in [3,4,5]:
        fig, axes = plt.subplots(3, 2, figsize=(14, 12))
        fig.suptitle(f"QNodes vs GeometricSIA — k={k}", fontsize=14)
        for idx, sheet in enumerate(SHEETS):
            tests = datos[sheet]
            labels = [f"T{i+1}" for i in range(len(tests))]
            q_phis = [t[f"Q{k}_phi"] if t[f"Q{k}_phi"] is not None else 0 for t in tests]
            g_phis = [t[f"G{k}_phi"] if t[f"G{k}_phi"] is not None else 0 for t in tests]
            q_times = [t[f"Q{k}_time"] if t[f"Q{k}_time"] is not None else 0 for t in tests]
            g_times = [t[f"G{k}_time"] if t[f"G{k}_time"] is not None else 0 for t in tests]

            x = np.arange(len(labels))
            w = 0.35
            ax = axes[idx, 0]
            ax.bar(x - w/2, q_phis, w, label="QNodes", color="#1f77b4")
            ax.bar(x + w/2, g_phis, w, label="GeometricSIA", color="#ff7f0e")
            ax.set_title(f"{sheet} — φ")
            ax.set_xticks(x)
            ax.set_xticklabels(labels, fontsize=8)
            ax.legend(fontsize=8)
            ax.set_ylabel("φ")

            ax = axes[idx, 1]
            ax.bar(x - w/2, q_times, w, label="QNodes", color="#1f77b4")
            ax.bar(x + w/2, g_times, w, label="GeometricSIA", color="#ff7f0e")
            ax.set_title(f"{sheet} — Tiempo (s)")
            ax.set_xticks(x)
            ax.set_xticklabels(labels, fontsize=8)
            ax.legend(fontsize=8)
            ax.set_ylabel("Tiempo (s)")

        plt.tight_layout()
        plt.savefig(str(REPORTS_DIR / f"comparacion_k{k}.png"), dpi=150)
        plt.close()
        print(f"  Grafico guardado: reports/comparacion_k{k}.png")

    fig, axes = plt.subplots(2, 2, figsize=(12, 10))
    fig.suptitle("Resumen por hoja", fontsize=14)
    for idx, sheet in enumerate(SHEETS):
        tests = datos[sheet]
        ks = [3,4,5]
        q_means = [np.mean([t[f"Q{k}_phi"] for t in tests if t[f"Q{k}_phi"] is not None]) for k in ks]
        g_means = [np.mean([t[f"G{k}_phi"] for t in tests if t[f"G{k}_phi"] is not None]) for k in ks]
        q_tmeans = [np.mean([t[f"Q{k}_time"] for t in tests if t[f"Q{k}_time"] is not None]) for k in ks]
        g_tmeans = [np.mean([t[f"G{k}_time"] for t in tests if t[f"G{k}_time"] is not None]) for k in ks]

        row, col = idx // 2, idx % 2
        ax = axes[row, col]
        x = np.arange(len(ks))
        w = 0.35
        ax.bar(x - w/2, q_means, w, label="QNodes", color="#1f77b4")
        ax.bar(x + w/2, g_means, w, label="GeometricSIA", color="#ff7f0e")
        ax.set_title(f"{sheet}")
        ax.set_xticks(x)
        ax.set_xticklabels([f"k={k}" for k in ks])
        ax.legend(fontsize=8)
        ax.set_ylabel("φ promedio")

    plt.tight_layout()
    plt.savefig(str(REPORTS_DIR / "resumen_phi_por_hoja.png"), dpi=150)
    plt.close()
    print(f"  Grafico guardado: reports/resumen_phi_por_hoja.png")

def main():
    print("Extrayendo datos del Excel...")
    datos = extraer_datos()

    print("\n=== RESUMEN POR HOJA ===")
    resumen_por_hoja(datos)

    print("\n=== COMPARACION POR TEST ===")
    comparar_por_test(datos)

    print("\n=== GENERANDO GRAFICAS ===")
    graficar_comparacion(datos)

    json_path = REPORTS_DIR / "datos_extraidos.json"
    with open(json_path, "w") as f:
        json.dump({s: [{k: v for k, v in t.items() if v != ""} for t in tests] for s, tests in datos.items()}, f, indent=2)
    print(f"\n  Datos extraidos: {json_path}")
    print(f"\nReportes guardados en: {REPORTS_DIR}/")

if __name__ == "__main__":
    main()
