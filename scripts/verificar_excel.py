#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Verifica que todas las celdas E6:F15 de cada hoja tengan valores.
Uso: uv run python scripts/verificar_excel.py
"""

import sys, io
from pathlib import Path

sys.stdout = io.TextIOWrapper(
    io.BufferedWriter(sys.stdout.buffer), encoding="utf-8", errors="replace"
)

import openpyxl

EXCEL_FILE = Path(__file__).parent.parent / "DatosPruebas2026_1 (1).xlsx"
SHEETS = ["15B-Elementos", "20A-Elementos", "22A-Elementos", "25A-Elementos "]

print("=" * 60)
print("  VERIFICACION DE RESULTADOS EN EXCEL")
print("=" * 60)

try:
    wb = openpyxl.load_workbook(str(EXCEL_FILE), data_only=True)
except FileNotFoundError:
    print(f"ERROR: No se encuentra {EXCEL_FILE}")
    sys.exit(1)

gran_ok = 0
gran_total = 0

for sheet_name in SHEETS:
    if sheet_name not in wb.sheetnames:
        print(f"\n{sheet_name:<20} HOJA NO ENCONTRADA")
        continue

    ws = wb[sheet_name]
    ok = 0
    total = 0

    for fila in range(6, 16):
        b = ws[f"B{fila}"].value
        c = ws[f"C{fila}"].value
        if not b or not c:
            continue
        total += 1
        phi = ws[f"E{fila}"].value
        tiempo = ws[f"F{fila}"].value
        if phi is not None and tiempo is not None:
            ok += 1
        estado = "OK" if (phi is not None and tiempo is not None) else "VACIO"
        print(
            f"  {sheet_name:<20} fila {fila:<3}  "
            f"E={str(phi):>12}  F={str(tiempo):>12}  [{estado}]"
        )

    print(f"  -> {sheet_name}: {ok}/{total} completadas\n")
    gran_ok += ok
    gran_total += total

wb.close()
print(f"  TOTAL: {gran_ok}/{gran_total} pruebas completadas")
