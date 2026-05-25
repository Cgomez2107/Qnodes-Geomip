#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Ejecuta QNodes para una hoja del Excel y escribe resultados
en las columnas D (Particion), E (Perdida), F (Tiempo).

Muestra en consola:
  - Output completo de QNodes
  - Los valores exactos que se insertan en cada celda del Excel

Uso:
    uv run --directory QNodes python "$(pwd)/scripts/run_experiment.py" --sheet "10A-Elementos"
    uv run --directory QNodes python "$(pwd)/scripts/run_experiment.py" --all
"""

import sys
import io
import os
import re
import time
import shutil
import argparse
from pathlib import Path

sys.stdout = io.TextIOWrapper(
    io.BufferedWriter(sys.stdout.buffer), encoding="utf-8", errors="replace"
)

SCRIPTS_DIR = Path(__file__).parent
PROYECTO = SCRIPTS_DIR.parent
sys.path.insert(0, str(SCRIPTS_DIR))

import subprocess
import openpyxl
from openpyxl.styles import PatternFill
from utils.converter import excel_a_bits
from utils.parser import parsear_salida_qnodes


# -------------------------------------------------------------------
# Constantes
# -------------------------------------------------------------------
EXCEL_FILE = PROYECTO / "DatosPruebas2026_1 (1).xlsx"
EXEC_PY = PROYECTO / "QNodes" / "exec.py"
MAIN_PY = PROYECTO / "QNodes" / "src" / "main.py"
QNODES_DIR = PROYECTO / "QNodes"
N_PRUEBAS = 10
TIMEOUT = 28800

SHEETS = {
    "10A-Elementos": {"n_nodos": 10, "pagina": "A"},
    "15B-Elementos": {"n_nodos": 15, "pagina": "B"},
    "20A-Elementos": {"n_nodos": 20, "pagina": "A"},
    "22A-Elementos": {"n_nodos": 22, "pagina": "A"},
    "25A-Elementos ": {"n_nodos": 25, "pagina": "A"},  # trailing space in Excel
}

# -------------------------------------------------------------------
# Modificacion de archivos
# -------------------------------------------------------------------

def setear_pagina_en_exec(pagina: str):
    with open(EXEC_PY, "r", encoding="utf-8") as f:
        content = f.read()
    content = re.sub(
        r'set_pagina_red_muestra\s*\(\s*"[A-Z]"\s*\)',
        f'set_pagina_red_muestra("{pagina}")',
        content,
    )
    with open(EXEC_PY, "w", encoding="utf-8") as f:
        f.write(content)


def setear_parametros_en_main(estado_inicial, condiciones, alcance, mecanismo):
    with open(MAIN_PY, "r", encoding="utf-8") as f:
        content = f.read()
    content = re.sub(
        r'estado_inicial\s*=\s*"[^"]*"',
        f'estado_inicial = "{estado_inicial}"',
        content,
    )
    content = re.sub(
        r'condiciones\s*=\s*"[^"]*"',
        f'condiciones = "{condiciones}"',
        content,
    )
    content = re.sub(
        r'alcance\s*=\s*"[^"]*"',
        f'alcance = "{alcance}"',
        content,
    )
    content = re.sub(
        r'mecanismo\s*=\s*"[^"]*"',
        f'mecanismo = "{mecanismo}"',
        content,
    )
    with open(MAIN_PY, "w", encoding="utf-8") as f:
        f.write(content)


def ejecutar_qnodes():
    try:
        result = subprocess.run(
            ["uv", "run", "exec.py"],
            cwd=str(QNODES_DIR),
            capture_output=True,
            text=True,
            encoding="utf-8",
            errors="replace",
            timeout=TIMEOUT,
        )
        return result.stdout + result.stderr
    except subprocess.TimeoutExpired:
        return None
    except Exception as e:
        return None


def tpm_existe(n_nodos: int, pagina: str) -> bool:
    return (QNODES_DIR / "src" / ".samples" / f"N{n_nodos}{pagina}.csv").exists()


def limpiar_particion(texto: str) -> str:
    """Elimina codigos ANSI y normaliza espacios de la particion."""
    if not texto:
        return ""
    # Eliminar secuencias ESC (ANSI)
    texto = re.sub(r'\x1b\[[0-9;]*[a-zA-Z]', '', texto)
    # Eliminar caracteres de control excepto newline
    texto = re.sub(r'[\x00-\x08\x0b\x0c\x0e-\x1f]', '', texto)
    # Normalizar espacios multiples
    texto = re.sub(r' {2,}', ' ', texto)
    return texto.strip()


# -------------------------------------------------------------------
# Procesamiento de una hoja
# -------------------------------------------------------------------

def procesar_hoja(sheet_name: str, config: dict) -> list:
    """Return list of results dicts for later summary."""
    n_nodos = config["n_nodos"]
    pagina = config["pagina"]

    print(f"\n{'='*70}")
    print(f"  {sheet_name} ({n_nodos} nodos, pagina {pagina})")
    print(f"{'='*70}")

    if not tpm_existe(n_nodos, pagina):
        print(f"  ERROR: Falta N{n_nodos}{pagina}.csv")
        return []

    print(f"\n  --- Configurando pagina {pagina} ---")
    setear_pagina_en_exec(pagina)
    print(f"  exec.py -> set_pagina_red_muestra(\"{pagina}\")")

    print(f"\n  --- Leyendo {N_PRUEBAS} pruebas desde Excel ---")
    try:
        wb = openpyxl.load_workbook(str(EXCEL_FILE), data_only=False)
        ws = wb[sheet_name]
        pruebas = []
        for fila in range(6, 6 + N_PRUEBAS):
            alcance = ws[f"B{fila}"].value
            mecanismo = ws[f"C{fila}"].value
            if alcance and mecanismo:
                pruebas.append({
                    "fila": fila,
                    "alcance": str(alcance).strip(),
                    "mecanismo": str(mecanismo).strip(),
                })
        wb.close()
        print(f"  Leidas {len(pruebas)} pruebas")
    except Exception as e:
        print(f"  ERROR: {e}")
        return []

    resumen = []
    print(f"\n  --- Ejecutando QNodes ---")
    for idx, p in enumerate(pruebas, 1):
        alcance_bits = excel_a_bits(p["alcance"], n_nodos)
        mecanismo_bits = excel_a_bits(p["mecanismo"], n_nodos)
        estado_inicial = "1" + "0" * (n_nodos - 1)
        condiciones = "1" * n_nodos

        setear_parametros_en_main(estado_inicial, condiciones, alcance_bits, mecanismo_bits)

        print(f"\n  >>> PRUEBA {idx}/{N_PRUEBAS}  (fila {p['fila']}) <<<")
        print(f"      alcance={p['alcance']} -> bits={alcance_bits}")
        print(f"      mecanismo={p['mecanismo']} -> bits={mecanismo_bits}")
        print(f"      estado_inicial={estado_inicial}  condiciones={condiciones}")
        print(f"      Ejecutando QNodes...")

        t_start = time.time()
        raw_output = ejecutar_qnodes()
        t_end = time.time()
        wall_time = t_end - t_start

        if raw_output is None:
            print(f"      [TIMEOUT] No se recibio respuesta en {TIMEOUT}s")
            continue

        print(f"\n      --- SALIDA CRUDA DE QNODES ---")
        for line in raw_output.strip().split("\n"):
            print(f"      | {line}")
        print(f"      --- FIN SALIDA ---")

        parsed = parsear_salida_qnodes(raw_output)
        phi = parsed.get("phi")
        particion = parsed.get("particion")
        tiempo = parsed.get("tiempo")

        if phi is None:
            print(f"      [ERROR] No se pudo extraer phi de la salida")
            continue

        # Limpiar ANSI de la particion para Excel
        particion_limpia = limpiar_particion(particion)

        print(f"\n      --- RESULTADO PARSEADO ---")
        print(f"      Particion (cruda) = {particion}")
        print(f"      Particion (limpia)= {particion_limpia}")
        print(f"      Phi       = {phi}")
        print(f"      Tiempo    = {tiempo} s  (wall: {wall_time:.4f}s)")

        res = {
            "fila": p["fila"],
            "alcance_excel": p["alcance"],
            "mecanismo_excel": p["mecanismo"],
            "particion": particion_limpia,
            "phi": phi,
            "tiempo": tiempo,
        }
        resumen.append(res)

        print(f"\n      --- ESCRIBIENDO EN EXCEL ---")
        print(f"      -> {EXCEL_FILE.name} / {sheet_name}")
        print(f"      -> Celda D{p['fila']} (Particion) = \"{particion}\"")
        print(f"      -> Celda E{p['fila']} (Perdida)   = {phi}")
        print(f"      -> Celda F{p['fila']} (Tiempo)    = {tiempo}")
        print(f"      {'='*50}")

    if not resumen:
        print(f"\n  No se obtuvieron resultados para {sheet_name}")
        return []

    print(f"\n  --- ESCRIBIENDO RESULTADOS EN EXCEL ---")
    exito = 0
    for r in resumen:
        try:
            time.sleep(0.5)
            wb = openpyxl.load_workbook(str(EXCEL_FILE))
            ws = wb[sheet_name]
            ws[f"D{r['fila']}"] = str(r['particion']) if r['particion'] else ""
            ws[f"E{r['fila']}"] = float(r['phi']) if r['phi'] is not None else ""
            ws[f"F{r['fila']}"] = float(r['tiempo']) if r['tiempo'] is not None else ""
            fill = PatternFill(start_color="FFB6D9EC", end_color="FFB6D9EC", fill_type="solid")
            for col in ["D", "E", "F"]:
                ws[f"{col}{r['fila']}"].fill = fill
            wb.save(str(EXCEL_FILE))
            wb.close()
            exito += 1
            print(f"      Fila {r['fila']}: OK")
        except Exception as e:
            print(f"      Fila {r['fila']}: ERROR {type(e).__name__}: {e}")

    print(f"  {exito}/{len(resumen)} resultados escritos correctamente")

    return resumen


# -------------------------------------------------------------------
# CLI
# -------------------------------------------------------------------

def main():
    parser = argparse.ArgumentParser(
        description="Ejecuta QNodes y escribe D=Particion, E=Perdida, F=Tiempo en Excel"
    )
    group = parser.add_mutually_exclusive_group(required=True)
    group.add_argument("--sheet", type=str, help="Nombre de la hoja")
    group.add_argument("--all", action="store_true", help="Todas las hojas")
    args = parser.parse_args()

    print("=" * 70)
    print("  PRUEBAS QNODES - BIPARTICIONES")
    print("  Columnas: D=Particion  E=Perdida  F=Tiempo")
    print("=" * 70)

    backup = PROYECTO / f"BACKUP_{Path(EXCEL_FILE).name}"
    if not backup.exists():
        shutil.copy(str(EXCEL_FILE), str(backup))
        print(f"\n  Backup creado: {backup.name}")

    if args.all:
        hojas = list(SHEETS.items())
    else:
        sheet_name = args.sheet
        if sheet_name not in SHEETS:
            print(f"ERROR: Hoja '{sheet_name}' no reconocida.")
            print(f"Opciones: {list(SHEETS.keys())}")
            sys.exit(1)
        hojas = [(sheet_name, SHEETS[sheet_name])]

    todos_resultados = []
    for sheet_name, config in hojas:
        res = procesar_hoja(sheet_name, config)
        todos_resultados.extend(res)

    # Resumen final
    print(f"\n{'='*70}")
    print(f"  RESUMEN FINAL - TODOS LOS RESULTADOS")
    print(f"{'='*70}")
    for r in todos_resultados:
        print(f"  {r['fila']:>3}  D=\"{r['particion']}\"  E={r['phi']}  F={r['tiempo']}")
    print(f"{'='*70}")
    print(f"  Total: {len(todos_resultados)} pruebas completadas")
    print(f"{'='*70}")


if __name__ == "__main__":
    main()
